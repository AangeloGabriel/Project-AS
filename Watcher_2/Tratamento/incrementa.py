from openpyxl import load_workbook
from datetime import date, datetime

#Contador de quantas linhas tem no arquivo
def ContaLinha(sheet):
    max_row = 0
    for i, column in enumerate(sheet.iter_rows(values_only=True),start=0):
        max_row += 1
    max = max_row
    return max_row

def ContaColuna(sheet):
    max_column = 0
    for i, column in enumerate(sheet.iter_cols(values_only=True),start=0):
        max_column += 1
    return max_column

def incrementa(arquivocru, oficial):
    #Calcula o dia atual e pega o numero da semana conforme a ISO 8601
    dia = date.today()
    data_atual = dia.strftime('%d/%m/%Y')
    semana_atual = dia.isocalendar()[1]
    
    #Carrega os arquivos
    Wb_Cru = load_workbook(arquivocru)
    Wb_Base = load_workbook(oficial)
    
    #Ativa as planilhas
    ws_cru = Wb_Cru.active
    ws_base = Wb_Base.active

    #Pega coluna e linha do arquivo Cru
    LinhaMaxCru = ContaLinha(ws_cru)
    ColunaMaxCru = ContaColuna(ws_cru)

    #Pega coluna e linha do arquivo Base
    LinhaMaxBase = ContaLinha(ws_base)
    ColunaMaxBase = ContaColuna(ws_base)
    
    linhasBase = LinhaMaxBase
    for row in ws_cru.iter_rows(min_row=2):
        valoreslinha = [cell.value for cell in row]
        # coloca data e semana nas colunas 1 e 2:
        ws_base.cell(row=linhasBase, column=1).value = data_atual
        ws_base.cell(row=linhasBase, column=2).value = semana_atual
        # coloca os dados a partir da coluna 3:
        for col_index, valor in enumerate(valoreslinha):
            ws_base.cell(row=linhasBase, column=col_index + 3).value = valor
        linhasBase += 1
    
    return Wb_Base