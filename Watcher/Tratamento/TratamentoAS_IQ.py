from openpyxl import load_workbook

def tratamento(arquivo1, arquivo2):

    wb_Brasil = load_workbook(arquivo1)
    wb_IQ = load_workbook(arquivo2)

    ws_IQ = wb_IQ.active

    ws_IQ.delete_cols(1)
    ws_IQ.delete_rows(1, 4)

    first_columns = 5
    columns_to_delet = []
    lista = []
    max_columns = []

    for i, value in enumerate(ws_IQ.columns):
        lista.append(i)
        max_columns = max(lista)

    max_columns = max_columns - 6

    for i in lista:
        if i > first_columns:
            if i < max_columns:
                columns_to_delet.append(i)
        else:
            None

    for i in sorted(columns_to_delet, reverse=True):
        ws_IQ.delete_cols(i + 1)
    
    ws_Brasil = wb_Brasil.active

    ws_Brasil.delete_cols(1)
    ws_Brasil.delete_rows(1, 3)
    
    first_columns = 5
    columns_to_delet = []
    lista = []
    max_columns = []

    for i, valor in enumerate(ws_Brasil.columns):
        lista.append(i)
    
    max_columns = max(lista)

    max_columns = max_columns - 7

    for i in lista:
        if i > first_columns:
            if i < max_columns:
                columns_to_delet.append(i+1)
        else:
            None

    for i in sorted(columns_to_delet, reverse=True):
        ws_Brasil.delete_cols(i)

    max_rows = 0

    for row in enumerate(ws_Brasil.rows):
        max_rows += 1

    for i, row in enumerate(ws_IQ.iter_rows()):
        for j, cell in enumerate(row):
            ws_Brasil.cell(row=max_rows + i + 1, column=j + 1).value = cell.value

    colunas = []
    for i, valor in enumerate(ws_Brasil.columns):
        colunas.append(i)
    
    for i in sorted(colunas, reverse=True): 
        if i >= 9: 
            ws_Brasil.delete_cols(i+1)
        else:
            continue
        
    return wb_Brasil